#!/usr/bin/env python3
"""
build_ems_dashboard.py - generator for the EMS operations dashboard (ems_financial_live.html)
from SmartBuilder exports + bank balances. Sibling of build_dashboards.py, but for the EMS
operations layer (not the SA Minerals investment layer).

It reads a LOCAL folder of files (synced from Google Drive 02_EMS/SmartBuilder + 02_EMS/*),
so it runs in an environment where those files are available as regular files
(Google Drive for Desktop / downloaded exports).

Expected files in --data (matched by keywords, newest is picked):
  * "Creditors Age Analysis*.xlsx"   - AP (payables), aging
  * "Debtors Age Analysis*.xlsx"     - AR (receivables), aging
  * "Transaction List*.xlsx"         - GL transactions (full FY2026) -> cost structure
  * ems_balances.json                - EMS bank balances + loan (see below)
  * ems_budget.json  (optional)      - {overheads_monthly, sars_arrears, target_ebitda, notes}

ems_balances.json (example):
  {"fnb_ems":0,"bidvest":907267,"absa":9408,"drilltec":1399293,"loan":8114083,"asof":"2026-06-24"}

Run:
  pip install openpyxl --break-system-packages
  python3 build_ems_dashboard.py --data <folder> --out ems_financial_live.html

Definitions (transparent, edit as needed):
  gross_obligations = AP_total + sars_arrears
  net_exposure      = gross_obligations - AR_total - cash_on_hand
  cash_on_hand      = sum of EMS bank balances
  AP/AR totals are shown EXCLUDING take-on (opening/migration) balances.
"""
import sys, os, json, glob, argparse, datetime, re

def _norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())

def newest(folder, *keywords):
    kws = [k.lower() for k in keywords]
    best, bestm = None, -1
    for fp in glob.glob(os.path.join(folder, "*.xlsx")):
        name = os.path.basename(fp).lower()
        if all(k in name for k in kws):
            m = os.path.getmtime(fp)
            if m > bestm:
                best, bestm = fp, m
    return best

def read_rows(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows

def find_header(rows, must_have):
    """Return (idx, {norm_colname: col_index}) for the header row."""
    want = [_norm(x) for x in must_have]
    for i, r in enumerate(rows[:40]):
        norm = [_norm(c) for c in r]
        if all(any(w == n or w in n for n in norm) for w in want):
            colmap = {}
            for j, c in enumerate(r):
                nc = _norm(c)
                if nc:
                    colmap.setdefault(nc, j)
            return i, colmap
    return None, None

def col(colmap, *aliases):
    for a in aliases:
        na = _norm(a)
        for k, v in colmap.items():
            if k == na or na in k:
                return v
    return None

def num(v):
    try:
        return round(float(str(v).replace(",", "").replace("(", "-").replace(")", "")), 2)
    except (ValueError, TypeError):
        return None

def parse_aging(path, kind):
    """kind: 'creditors'|'debtors'. Returns dict with total, total_ex, buckets, top, takeon."""
    rows = read_rows(path)
    hidx, cm = find_header(rows, ["Name", "Totals"])
    if hidx is None:
        return {"total": 0.0, "total_ex": 0.0, "buckets": {}, "top": [], "takeon": 0.0, "n": 0, "error": "header not found"}
    c_name = col(cm, "name")
    c_code = col(cm, "code")
    c_tot  = col(cm, "totals", "total")
    c_cur  = col(cm, "current")
    c_30   = col(cm, "30days")
    c_60   = col(cm, "60days")
    c_90   = col(cm, "90days")
    # ">90 Days" is a separate column after "90 Days"
    c_over = None
    for k, v in cm.items():
        if k.startswith("90days") and v != c_90:
            c_over = v
    total = 0.0; takeon = 0.0; n = 0
    buckets = {"current": 0.0, "d30": 0.0, "d60": 0.0, "d90": 0.0, "d90p": 0.0}
    items = []
    for r in rows[hidx + 1:]:
        if c_code is None or c_code >= len(r):
            continue
        code = r[c_code]
        name = r[c_name] if c_name is not None and c_name < len(r) else None
        t = num(r[c_tot]) if c_tot is not None and c_tot < len(r) else None
        if not code or not name or t is None:
            continue
        if str(name).lower().startswith("category"):
            continue
        total += t; n += 1
        if "take on" in str(name).lower() or "take-on" in str(name).lower():
            takeon += t
        for key, ci in (("current", c_cur), ("d30", c_30), ("d60", c_60), ("d90", c_90), ("d90p", c_over)):
            if ci is not None and ci < len(r):
                b = num(r[ci])
                if b:
                    buckets[key] += b
        items.append((str(name)[:38], t))
    items.sort(key=lambda x: abs(x[1]), reverse=True)
    return {"total": round(total, 2), "total_ex": round(total - takeon, 2),
            "buckets": {k: round(v, 2) for k, v in buckets.items()},
            "top": items[:8], "takeon": round(takeon, 2), "n": n}

def parse_transactions(path):
    """Cost structure by GL: group by ledger name, P&L only (code not 9xxxxx)."""
    rows = read_rows(path)
    hidx, cm = find_header(rows, ["LEDGERCODESLEDGERNAME", "DEBIT", "CREDIT"])
    if hidx is None:
        return {"total": 0.0, "by_cat": [], "n": 0, "error": "header not found"}
    c_lname = col(cm, "ledgercodesledgername")
    c_lcode = col(cm, "transactionsledgercode")
    c_deb   = col(cm, "debit")
    c_cred  = col(cm, "credit")
    c_home  = col(cm, "homecurramount")
    cats = {}; n = 0
    for r in rows[hidx + 1:]:
        if c_lcode is None or c_lcode >= len(r):
            continue
        code = str(r[c_lcode] or "").strip()
        if not code or code.startswith("9"):   # 9xxxxx = balance sheet, not P&L
            continue
        name = r[c_lname] if c_lname is not None and c_lname < len(r) else None
        if not name:
            continue
        amt = num(r[c_home]) if c_home is not None and c_home < len(r) else None
        if amt is None:
            d = num(r[c_deb]) or 0.0; c = num(r[c_cred]) or 0.0
            amt = d - c
        cats[str(name)] = cats.get(str(name), 0.0) + amt
        n += 1
    by = sorted(((k, round(v, 2)) for k, v in cats.items()), key=lambda x: abs(x[1]), reverse=True)
    total = round(sum(v for _, v in by if v > 0), 2)
    return {"total": total, "by_cat": by[:12], "n": n}

def load_json(path):
    try:
        return json.load(open(path, encoding="utf-8"))
    except (OSError, ValueError):
        return {}

def M(n):
    return f"R{n/1e6:.2f}m"

def rm(n):
    return f"{n:,.0f}"

CSS = """*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Georgia,serif;background:#F5F2EA;color:#1A1A1A;line-height:1.5}
.wrap{max-width:1080px;margin:0 auto;padding:28px 22px 70px}
h1{font-size:25px}.top{font-size:10px;letter-spacing:3px;text-transform:uppercase;color:#8A8A82}
.hr{height:1px;background:#1A1A1A;border:none;margin:16px 0 20px}
.kgrid{display:grid;grid-template-columns:repeat(4,1fr);gap:1px;background:#D9D4C7;border:1px solid #D9D4C7;margin:20px 0}
.k{background:#FCFAF4;padding:15px 16px}.k .l{font-size:9.5px;letter-spacing:1px;text-transform:uppercase;color:#8A8A82}
.k .v{font-size:22px;font-weight:700;margin-top:5px}.k .m{font-size:10.5px;color:#555;margin-top:3px}
.sec{font-size:11px;letter-spacing:2px;text-transform:uppercase;border-bottom:1px solid #1A1A1A;padding-bottom:6px;margin:30px 0 14px;font-weight:700}
.sec span{float:right;font-size:9.5px;color:#8A8A82;text-transform:none;font-weight:400}
table{width:100%;border-collapse:collapse;font-size:12.5px}
th{text-align:left;font-size:9.5px;letter-spacing:.8px;text-transform:uppercase;color:#8A8A82;padding:7px 9px;border-bottom:1px solid #1A1A1A}
td{padding:7px 9px;border-bottom:1px solid #D9D4C7}td.n,th.n{text-align:right;font-variant-numeric:tabular-nums}
tr.tot td{font-weight:700;border-top:1px solid #1A1A1A}
.brow{display:flex;align-items:center;gap:10px;margin:6px 0}.blab{width:150px;font-size:12px;flex-shrink:0}
.btr{flex:1;background:#EAE5D8;height:15px;border-radius:1px;overflow:hidden}.bfl{height:100%;background:#1A1A1A}
.bamt{width:120px;text-align:right;font-size:12px;font-variant-numeric:tabular-nums}
.foot{font-size:10.5px;color:#8A8A82;margin-top:30px;border-top:1px solid #D9D4C7;padding-top:12px}
@media(max-width:760px){.kgrid{grid-template-columns:1fr 1fr}}"""

def aging_table(a, label):
    b = a["buckets"]
    rows = "".join(f"<tr><td>{i+1:02d}</td><td>{n}</td><td class='n'>{rm(v)}</td></tr>"
                   for i, (n, v) in enumerate(a["top"]))
    takeon_note = (f" &middot; take-on {rm(a['takeon'])} &rarr; gross {rm(a['total'])}" if a['takeon'] else "")
    return f"""<div class="sec">{label} <span>{a['n']} accounts &middot; top-8</span></div>
<table><tr><th>#</th><th>Counterparty</th><th class="n">Balance (ZAR)</th></tr>{rows}
<tr class="tot"><td></td><td>TOTAL (excl. take-on)</td><td class="n">{rm(a['total_ex'])}</td></tr></table>
<div style="font-size:11px;color:#555;margin-top:6px">Current {rm(b.get('current',0))} &middot; 30d {rm(b.get('d30',0))} &middot; 60d {rm(b.get('d60',0))} &middot; 90d {rm(b.get('d90',0))} &middot; &gt;90d {rm(b.get('d90p',0))}{takeon_note}</div>"""

def render(bal, ap, ar, tx, budget, gen):
    cash = sum(float(bal.get(k, 0) or 0) for k in ("fnb_ems", "bidvest", "absa", "drilltec"))
    loan = float(bal.get("loan", 0) or 0)
    paye = float(budget.get("paye_arrears", 0) or 0)
    vat  = float(budget.get("vat_arrears", 0) or 0)
    sars = float(budget.get("sars_arrears", 0) or 0) or (paye + vat)
    gross = ap["total_ex"] + sars
    net = gross - ar["total_ex"] - cash
    # Obligations & SARS exposure (SARS is a manual input via ems_budget.json)
    sars_rows = ""
    if paye or vat:
        sars_rows = (f'<tr><td>SARS &ndash; PAYE arrears</td><td class="n">{rm(paye)}</td></tr>'
                     f'<tr><td>SARS &ndash; VAT arrears</td><td class="n">{rm(vat)}</td></tr>')
    elif sars:
        sars_rows = f'<tr><td>SARS arrears</td><td class="n">{rm(sars)}</td></tr>'
    else:
        sars_rows = '<tr><td>SARS arrears <i>(add to ems_budget.json)</i></td><td class="n">0</td></tr>'
    obl_html = f"""<div class="sec">Obligations &amp; SARS exposure</div>
<table><tr><th>Item</th><th class="n">ZAR</th></tr>
<tr><td>AP &ndash; suppliers (excl. take-on)</td><td class="n">{rm(ap['total_ex'])}</td></tr>
{sars_rows}
<tr class="tot"><td>Gross obligations</td><td class="n">{rm(gross)}</td></tr>
<tr><td>less AR (excl. take-on)</td><td class="n">{rm(ar['total_ex'])}</td></tr>
<tr><td>less cash on hand</td><td class="n">{rm(cash)}</td></tr>
<tr class="tot"><td>Net exposure</td><td class="n">{rm(net)}</td></tr></table>"""
    # Day Zero - frozen baseline (manual, not auto-rebuilt)
    dz = budget.get("day_zero") or {}
    dz_html = (f'<div class="sec">Day Zero &ndash; frozen baseline <span>fixed manually, not auto-rebuilt</span></div>'
               f'<div style="font-size:12.5px;color:#333">&#127916; <b>{dz.get("date","")}</b> &ndash; {dz.get("note","")}</div>'
               if dz else "")
    # Path to EBITDA+ (manual plan)
    p = budget.get("path") or {}
    path_html = (f"""<div class="sec">Path to EBITDA+ <span>{p.get('window','')}</span></div>
<div class="kgrid">
<div class="k"><div class="l">Revenue plan</div><div class="v">{M(float(p.get('revenue',0) or 0))}</div><div class="m">{p.get('window','')}</div></div>
<div class="k"><div class="l">EBITDA (year)</div><div class="v">{M(float(p.get('ebitda_year',0) or 0))}</div></div>
<div class="k"><div class="l">First EBITDA+</div><div class="v" style="font-size:16px">{p.get('first_positive','n/a')}</div></div>
<div class="k"><div class="l">Condition</div><div class="v" style="font-size:11px">{p.get('condition','')[:70]}</div></div>
</div>""" if p else "")
    tot_cost = sum(v for _, v in tx["by_cat"] if v > 0) or 1
    bars = "".join(
        f"<div class='brow'><div class='blab'>{n[:26]}</div><div class='btr'><div class='bfl' style='width:{max(v,0)/tot_cost*100:.1f}%'></div></div><div class='bamt'>{rm(v)}</div></div>"
        for n, v in tx["by_cat"] if v > 0)
    ovh = budget.get("overheads_monthly")
    ebitda = budget.get("target_ebitda")
    if ovh or ebitda:
        budget_html = f"""<div class="sec">Budget P&amp;L FY2026-2027</div>
<div class="kgrid"><div class="k"><div class="l">General Overheads /mo</div><div class="v">{M(float(ovh)) if ovh else 'n/a'}</div></div>
<div class="k"><div class="l">Target EBITDA /mo</div><div class="v">{M(float(ebitda)) if ebitda else 'n/a'}</div></div>
<div class="k"><div class="l">Notes</div><div class="v" style="font-size:12px">{budget.get('notes','')[:60]}</div></div>
<div class="k"><div class="l">Source</div><div class="v" style="font-size:12px">ems_budget.json</div></div></div>"""
    else:
        budget_html = '<div class="sec">Budget P&amp;L</div><div style="font-size:12px;color:#555">Add <b>ems_budget.json</b> {overheads_monthly, target_ebitda, sars_arrears, notes} to populate this block.</div>'
    return f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>EMS Mining - Operations Dashboard</title><style>{CSS}</style></head><body><div class="wrap">
<div class="top">Engineered Mining Solutions (Pty) Ltd &middot; operations layer</div>
<h1>EMS Mining - Operations</h1><hr class="hr">
<div class="kgrid">
<div class="k"><div class="l">EMS cash on hand</div><div class="v">{M(cash)}</div><div class="m">FNB+Bidvest+ABSA+DrillTec</div></div>
<div class="k"><div class="l">Loan received</div><div class="v">{M(loan)}</div><div class="m">from SA Minerals (drawdown)</div></div>
<div class="k"><div class="l">Gross obligations</div><div class="v">{M(gross)}</div><div class="m">AP + SARS arrears</div></div>
<div class="k"><div class="l">Net exposure</div><div class="v">{M(net)}</div><div class="m">obligations &minus; AR &minus; cash</div></div>
</div>
<div class="sec">EMS cash by bank account <span>as of {bal.get('asof','?')}</span></div>
<table><tr><th>Account</th><th class="n">Balance (ZAR)</th></tr>
<tr><td>FNB EMS Cheque</td><td class="n">{rm(float(bal.get('fnb_ems',0) or 0))}</td></tr>
<tr><td>Bidvest</td><td class="n">{rm(float(bal.get('bidvest',0) or 0))}</td></tr>
<tr><td>ABSA</td><td class="n">{rm(float(bal.get('absa',0) or 0))}</td></tr>
<tr><td>Drill Tec</td><td class="n">{rm(float(bal.get('drilltec',0) or 0))}</td></tr>
<tr class="tot"><td>Cash on hand</td><td class="n">{rm(cash)}</td></tr></table>
{obl_html}
{aging_table(ar, 'AR - clients owe us')}
{aging_table(ap, 'AP - we owe suppliers')}
<div class="sec">Cost structure - SmartBuilder GL <span>{tx['n']} P&amp;L postings &middot; top-12</span></div>
{bars if bars else '<div style="font-size:12px;color:#555">No P&L postings (check Transaction List).</div>'}
{budget_html}
{path_html}
{dz_html}
<div class="foot">Generated by build_ems_dashboard.py &middot; {gen}. Definitions: gross = AP + SARS; net = gross &minus; AR &minus; cash. AP/AR shown excluding take-on (opening) balances. Bank balances from ems_balances.json.</div>
</div></body></html>"""

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", required=True, help="folder with EMS exports (SmartBuilder + ems_balances.json)")
    ap.add_argument("--out", required=True, help="output HTML")
    args = ap.parse_args()
    d = args.data
    f_cred = newest(d, "creditors", "age")
    f_debt = newest(d, "debtors", "age")
    f_tx   = newest(d, "transaction", "list")
    bal    = load_json(os.path.join(d, "ems_balances.json"))
    budget = load_json(os.path.join(d, "ems_budget.json"))
    missing = [n for n, f in [("Creditors Age", f_cred), ("Debtors Age", f_debt), ("Transaction List", f_tx)] if not f]
    if missing:
        print("WARNING: files not found: " + ", ".join(missing), file=sys.stderr)
    ap_d = parse_aging(f_cred, "creditors") if f_cred else {"total":0,"total_ex":0,"buckets":{},"top":[],"takeon":0,"n":0}
    ar_d = parse_aging(f_debt, "debtors") if f_debt else {"total":0,"total_ex":0,"buckets":{},"top":[],"takeon":0,"n":0}
    tx_d = parse_transactions(f_tx) if f_tx else {"total":0,"by_cat":[],"n":0}
    gen = datetime.date.today().strftime("%d %b %Y")
    html = render(bal, ap_d, ar_d, tx_d, budget, gen)
    with open(args.out, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"OK EMS ops -> {args.out}")
    print(f"  cash={M(sum(float(bal.get(k,0) or 0) for k in ('fnb_ems','bidvest','absa','drilltec')))} "
          f"AP={M(ap_d['total_ex'])} AR={M(ar_d['total_ex'])} cost_lines={len(tx_d['by_cat'])} "
          f"files: cred={bool(f_cred)} debt={bool(f_debt)} tx={bool(f_tx)}")

if __name__ == "__main__":
    main()
